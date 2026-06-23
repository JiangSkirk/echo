"""Office document tools: Excel and PDF generation/manipulation."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from js.config import ToolLimits
from js.security.guard import BehaviorGuard, SecurityDecisionType
from js.tools.registry import ToolParam, ToolResult, ToolSpec

_OFFICE_EXTRA_MSG = "Install js-agent[office] to use Excel tools."
_PDF_EXTRA_MSG = "Install js-agent[pdf] to use PDF generation tools."


class OfficeTools:
    """Tools for Excel and PDF document operations."""

    def __init__(self, workspace: Path, limits: ToolLimits, guard: BehaviorGuard) -> None:
        self.workspace = workspace.resolve()
        self.limits = limits
        self.guard = guard

    def _resolve(self, path: str) -> Path:
        p = Path(path)
        if p.is_absolute():
            resolved = p.resolve()
        else:
            resolved = (self.workspace / p).resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as e:
            raise ValueError(f"Path escapes workspace: {path}") from e
        return resolved

    def get_specs(self) -> list[ToolSpec]:
        return [
            ToolSpec(
                name="csv_read",
                description="Read data from a CSV file. Returns JSON array of rows.",
                parameters=[
                    ToolParam("path", "string", "Path to CSV file (relative to workspace)"),
                    ToolParam(
                        "encoding", "string", "File encoding (default: utf-8)", required=False
                    ),
                    ToolParam(
                        "delimiter", "string", "Column delimiter (default: comma)", required=False
                    ),
                ],
                read_only=True,
            ),
            ToolSpec(
                name="csv_write",
                description="Write data to a CSV file.",
                parameters=[
                    ToolParam("path", "string", "Path to CSV file"),
                    ToolParam("data", "string", "JSON array of rows to write"),
                    ToolParam(
                        "encoding", "string", "File encoding (default: utf-8)", required=False
                    ),
                    ToolParam(
                        "delimiter", "string", "Column delimiter (default: comma)", required=False
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_read",
                description="Read data from an Excel file (.xlsx). Returns JSON array of rows.",
                parameters=[
                    ToolParam("path", "string", "Path to Excel file (relative to workspace)"),
                    ToolParam(
                        "sheet", "string", "Sheet name (default: first sheet)", required=False
                    ),
                    ToolParam("start_row", "integer", "1-based start row", required=False),
                    ToolParam("end_row", "integer", "1-based end row (inclusive)", required=False),
                    ToolParam(
                        "start_col", "string", "Start column letter (e.g. 'A')", required=False
                    ),
                    ToolParam("end_col", "string", "End column letter (e.g. 'Z')", required=False),
                ],
                read_only=True,
            ),
            ToolSpec(
                name="excel_write",
                description="Write data to an Excel file. Creates the file if it doesn't exist.",
                parameters=[
                    ToolParam("path", "string", "Path to Excel file"),
                    ToolParam("sheet", "string", "Sheet name (default: 'Sheet1')", required=False),
                    ToolParam(
                        "data", "string", 'JSON array of rows to write (e.g. [["A","B"],[1,2]])'
                    ),
                    ToolParam(
                        "start_cell",
                        "string",
                        "Start cell (e.g. 'A1', default: 'A1')",
                        required=False,
                    ),
                    ToolParam(
                        "append",
                        "boolean",
                        "Append to existing sheet instead of overwriting",
                        required=False,
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_merge",
                description=(
                    "Merge data from one Excel file into another at a specific location. "
                    "Useful for combining data: e.g. copy rows from source file and paste into target file at column E."
                ),
                parameters=[
                    ToolParam("source_path", "string", "Source Excel file path"),
                    ToolParam("target_path", "string", "Target Excel file path"),
                    ToolParam("source_sheet", "string", "Source sheet name", required=False),
                    ToolParam("target_sheet", "string", "Target sheet name", required=False),
                    ToolParam(
                        "source_range",
                        "string",
                        "Range like 'A1:D10' or leave empty for all data",
                        required=False,
                    ),
                    ToolParam(
                        "target_start_cell",
                        "string",
                        "Target start cell, e.g. 'E1' (default: 'A1')",
                        required=False,
                    ),
                    ToolParam(
                        "include_header",
                        "boolean",
                        "Include header row from source",
                        required=False,
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_create",
                description="Create a new blank Excel file with optional headers.",
                parameters=[
                    ToolParam("path", "string", "Output file path"),
                    ToolParam(
                        "sheet_name", "string", "Sheet name (default: 'Sheet1')", required=False
                    ),
                    ToolParam("headers", "string", "JSON array of column headers", required=False),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="pdf_generate",
                description="Generate a PDF file from tabular data.",
                parameters=[
                    ToolParam("path", "string", "Output PDF file path"),
                    ToolParam("title", "string", "Document title", required=False),
                    ToolParam("data", "string", "JSON array of rows (first row is header)"),
                    ToolParam("page_size", "string", "A4 or LETTER (default: A4)", required=False),
                ],
                dangerous=True,
            ),
        ]

    async def csv_read(
        self,
        path: str,
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "read")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        try:
            import csv

            if not target.exists():
                return ToolResult(success=False, error=f"File not found: {path}")

            rows: list[list[str]] = []
            with open(target, encoding=encoding, newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    rows.append(row)

            return ToolResult(
                success=True,
                output=json.dumps(rows, ensure_ascii=False, indent=2),
                metadata={"rows": len(rows), "columns": len(rows[0]) if rows else 0},
            )
        except UnicodeDecodeError:
            return ToolResult(
                success=False,
                error=f"Encoding error: {encoding} does not match file content. Try 'gbk' or 'latin-1'.",
            )
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    async def csv_write(
        self,
        path: str,
        data: str = "",
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "write")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        try:
            import csv

            rows_data: list[list[Any]] = json.loads(data) if data else []
            target.parent.mkdir(parents=True, exist_ok=True)
            with open(target, "w", encoding=encoding, newline="") as f:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerows(rows_data)

            return ToolResult(
                success=True,
                output=f"Written {len(rows_data)} rows to {path}",
                metadata={"path": str(target), "rows": len(rows_data)},
            )
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    @staticmethod
    def _parse_cell(value: Any) -> Any:
        """Preserve native types (int, float, bool) where possible."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        return str(value)

    async def excel_read(
        self,
        path: str,
        sheet: str | None = None,
        start_row: int = 0,
        end_row: int = 0,
        start_col: str = "",
        end_col: str = "",
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "read")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        wb = None
        try:
            try:
                from openpyxl import load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            if not target.exists():
                return ToolResult(success=False, error=f"File not found: {path}")

            wb = load_workbook(str(target), data_only=True, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                return ToolResult(success=False, error="Sheet not found")

            min_row = start_row if start_row > 0 else 1
            max_row = end_row if end_row > 0 else ws.max_row
            min_col = column_index_from_string(start_col) if start_col else 1
            max_col = column_index_from_string(end_col) if end_col else ws.max_column

            rows: list[list[Any]] = []
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                rows.append([self._parse_cell(cell) for cell in row])

            return ToolResult(
                success=True,
                output=json.dumps(rows, ensure_ascii=False, indent=2),
                metadata={"rows": len(rows), "columns": len(rows[0]) if rows else 0},
            )
        except KeyError as e:
            return ToolResult(success=False, error=f"Sheet not found: {e}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()

    @staticmethod
    def _parse_cell_ref(ref: str) -> tuple[str, str]:
        """Split Excel cell reference like 'A1' or 'BC123' into (col_letters, row_num)."""
        col = ""
        row = ""
        for ch in ref:
            if ch.isalpha():
                if row:
                    raise ValueError(f"Invalid cell reference: {ref}")
                col += ch
            elif ch.isdigit():
                row += ch
            else:
                raise ValueError(f"Invalid cell reference: {ref}")
        if not col or not row:
            raise ValueError(f"Invalid cell reference: {ref}")
        return col, row

    async def excel_write(
        self,
        path: str,
        sheet: str | None = None,
        data: str = "",
        start_cell: str = "A1",
        append: bool = False,
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "write")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        wb = None
        try:
            try:
                from openpyxl import Workbook, load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            sheet_name = sheet or "Sheet1"
            rows_data: list[list[Any]] = json.loads(data) if data else []
            if not isinstance(rows_data, list) or any(not isinstance(r, list) for r in rows_data):
                return ToolResult(success=False, error="data must be a JSON array of arrays")

            if target.exists():
                wb = load_workbook(str(target))
            else:
                wb = Workbook()
                if sheet_name != "Sheet1" and wb.active:
                    wb.active.title = sheet_name

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            col_letter, row_num_str = self._parse_cell_ref(start_cell)
            start_col = column_index_from_string(col_letter)
            start_row_num = int(row_num_str)

            if append:
                start_row_num = max(start_row_num, ws.max_row + 1)

            for r_idx, row in enumerate(rows_data, start=start_row_num):
                for c_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(target))

            return ToolResult(
                success=True,
                output=f"Written {len(rows_data)} rows to {path}",
                metadata={"path": str(target), "rows": len(rows_data)},
            )
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()

    async def excel_merge(
        self,
        source_path: str,
        target_path: str,
        source_sheet: str | None = None,
        target_sheet: str | None = None,
        source_range: str = "",
        target_start_cell: str = "A1",
        include_header: bool = True,
    ) -> ToolResult:
        try:
            source = self._resolve(source_path)
            target = self._resolve(target_path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        for p, op in [(source, "read"), (target, "write")]:
            decision = self.guard.check_path_operation(str(p), op)
            if decision.decision == SecurityDecisionType.BLOCK:
                return ToolResult(success=False, error=decision.reason)

        src_wb = None
        tgt_wb = None
        try:
            try:
                from openpyxl import load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string, range_boundaries
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            if not source.exists():
                return ToolResult(success=False, error=f"Source not found: {source_path}")
            if not target.exists():
                return ToolResult(success=False, error=f"Target not found: {target_path}")

            src_wb = load_workbook(str(source), data_only=True, read_only=True)
            src_ws = src_wb[source_sheet] if source_sheet else src_wb.active
            if src_ws is None:
                return ToolResult(success=False, error="Source sheet not found")

            tgt_wb = load_workbook(str(target))
            tgt_ws = tgt_wb[target_sheet] if target_sheet else tgt_wb.active
            if tgt_ws is None:
                return ToolResult(success=False, error="Target sheet not found")

            # Determine source data range
            if source_range:
                min_col, min_row, max_col, max_row = range_boundaries(source_range)
            else:
                min_row, min_col = 1, 1
                max_row = src_ws.max_row
                max_col = src_ws.max_column

            if not include_header and min_row == 1 and max_row is not None and max_row > 1:
                min_row += 1

            col_letter, row_num_str = self._parse_cell_ref(target_start_cell)
            tgt_start_col = column_index_from_string(col_letter)
            tgt_start_row = int(row_num_str)

            rows_copied = 0
            for r_idx, row in enumerate(
                src_ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                start=tgt_start_row,
            ):
                for c_idx, value in enumerate(row, start=tgt_start_col):
                    tgt_ws.cell(row=r_idx, column=c_idx, value=value)
                rows_copied += 1

            tgt_wb.save(str(target))

            return ToolResult(
                success=True,
                output=f"Merged {rows_copied} rows from {source_path} into {target_path} at {target_start_cell}",
                metadata={"rows_copied": rows_copied},
            )
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if src_wb is not None:
                src_wb.close()
            if tgt_wb is not None:
                tgt_wb.close()

    async def excel_create(
        self,
        path: str,
        sheet_name: str = "Sheet1",
        headers: str = "",
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "write")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        wb = None
        try:
            try:
                from openpyxl import Workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws.title = sheet_name

            if headers:
                hdrs: list[str] = json.loads(headers)
                if not isinstance(hdrs, list):
                    return ToolResult(success=False, error="headers must be a JSON array")
                for c_idx, h in enumerate(hdrs, start=1):
                    ws.cell(row=1, column=c_idx, value=h)

            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(target))

            return ToolResult(success=True, output=f"Created Excel file: {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()

    async def pdf_generate(
        self,
        path: str,
        title: str = "",
        data: str = "",
        page_size: str = "A4",
    ) -> ToolResult:
        try:
            target = self._resolve(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "write")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        try:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, LETTER
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import (
                    Paragraph,
                    SimpleDocTemplate,
                    Spacer,
                    Table,
                    TableStyle,
                )
            except ImportError as _e:
                raise ImportError(_PDF_EXTRA_MSG) from _e

            rows_data: list[list[Any]] = json.loads(data) if data else []
            if not rows_data:
                return ToolResult(success=False, error="No data provided")

            size = A4 if page_size.upper() == "A4" else LETTER
            target.parent.mkdir(parents=True, exist_ok=True)
            doc = SimpleDocTemplate(str(target), pagesize=size)

            elements: list[Any] = []
            styles = getSampleStyleSheet()

            if title:
                elements.append(Paragraph(title, styles["Title"]))
                elements.append(Spacer(1, 12))

            # Build table
            table = Table(rows_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ]
                )
            )
            elements.append(table)
            doc.build(elements)

            return ToolResult(success=True, output=f"Generated PDF: {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    def register_all(self, registry: Any) -> None:
        """Register all office tools."""
        for spec in self.get_specs():
            if spec.name == "csv_read":
                registry.register(spec, self.csv_read)
            elif spec.name == "csv_write":
                registry.register(spec, self.csv_write)
            elif spec.name == "excel_read":
                registry.register(spec, self.excel_read)
            elif spec.name == "excel_write":
                registry.register(spec, self.excel_write)
            elif spec.name == "excel_merge":
                registry.register(spec, self.excel_merge)
            elif spec.name == "excel_create":
                registry.register(spec, self.excel_create)
            elif spec.name == "pdf_generate":
                registry.register(spec, self.pdf_generate)
